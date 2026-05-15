from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.config import SOURCE_WORKBOOK, resolve_project_path
from app.database import SessionLocal, engine
from app.models import Base, Chapter, LearningDay, Module, QuizQuestion, Resource
from app.services.resource_catalog import ensure_resource_catalog

URL_RE = re.compile(r"(https?://[^\s）)]+)")

MODULE_RANGES = [
    ("0. 最小环境与学习仓库", 1, 3),
    ("1. Prompt / 结构化输出", 4, 7),
    ("2. LLM 基础 / Attention / Embedding", 8, 10),
    ("3. Function / Tool Calling", 11, 14),
    ("4. MCP", 15, 17),
    ("5. LangGraph / Agent 状态机", 18, 23),
    ("6. Memory / 长短期记忆", 24, 28),
    ("7. RAG / 知识库 Agent", 29, 34),
    ("8. Dify / Workflow 对照", 35, 37),
    ("9. Agent 工程 / OpenClaw / AI Coding / 毕业项目", 38, 50),
    ("10. 微调认知 / 数据集 / 训练 / 评估", 51, 65),
    ("11. 微调工程化 / 综合复盘", 66, 70),
]

CHAPTERS = [
    ("0. 最小环境与学习仓库", "理论基础篇", 1, 1, "理解目录、Git、API Key 安全和每天输出物的意义"),
    ("0. 最小环境与学习仓库", "进阶篇", 2, 2, "建立学习日志、问题清单和错误复盘模板"),
    ("0. 最小环境与学习仓库", "实战篇", 3, 3, "创建仓库和固定目录，准备后续所有实验"),
    ("1. Prompt / 结构化输出", "理论基础篇", 4, 4, "理解 prompt 的角色、目标、上下文、约束和格式"),
    ("1. Prompt / 结构化输出", "进阶篇", 5, 6, "掌握 few-shot、ReAct 风格、结构化输出和失败重试"),
    ("1. Prompt / 结构化输出", "实战篇", 7, 7, "用 eval 判断 prompt 是否真的更稳定"),
    ("2. LLM 基础 / Attention / Embedding", "理论基础篇", 8, 8, "理解 token、context、attention、QKV、embedding 的直觉"),
    ("2. LLM 基础 / Attention / Embedding", "进阶篇", 9, 9, "理解 embedding 相似度和检索边界"),
    ("2. LLM 基础 / Attention / Embedding", "实战篇", 10, 10, "把 LLM 基础和 Prompt 复盘成可用心智模型"),
    ("3. Function / Tool Calling", "理论基础篇", 11, 11, "理解 tool/function/schema/call/result 链路"),
    ("3. Function / Tool Calling", "进阶篇", 12, 13, "设计文件工具和只读 SQL 工具，处理错误和边界"),
    ("3. Function / Tool Calling", "实战篇", 14, 14, "总结工具安全、allowlist、日志和人工确认"),
    ("4. MCP", "理论基础篇", 15, 15, "理解 MCP 与普通 function calling 的区别"),
    ("4. MCP", "进阶篇", 16, 16, "设计文件 MCP 工具和资源暴露边界"),
    ("4. MCP", "实战篇", 17, 17, "完成多工具编排并记录调用日志"),
    ("5. LangGraph / Agent 状态机", "理论基础篇", 18, 19, "理解 state、node、edge、conditional edge"),
    ("5. LangGraph / Agent 状态机", "进阶篇", 20, 22, "掌握 planner/executor/reviewer、tool 节点和 human-in-the-loop"),
    ("5. LangGraph / Agent 状态机", "实战篇", 23, 23, "做一个最小 multi-agent 或多角色协作 demo"),
    ("6. Memory / 长短期记忆", "理论基础篇", 24, 24, "区分短期记忆、长期记忆、语义召回和用户画像"),
    ("6. Memory / 长短期记忆", "进阶篇", 25, 27, "实现 short-term、long-term、semantic recall 的读取与写入规则"),
    ("6. Memory / 长短期记忆", "实战篇", 28, 28, "复盘 memory 设计，决定毕业项目如何使用记忆"),
    ("7. RAG / 知识库 Agent", "理论基础篇", 29, 29, "理解 chunk、embedding、retrieval、rerank、answer generation"),
    ("7. RAG / 知识库 Agent", "进阶篇", 30, 33, "实验 chunking、hybrid/rerank、parent retrieval 和 RAG eval"),
    ("7. RAG / 知识库 Agent", "实战篇", 34, 34, "完成知识库 Agent v1，并用固定问题集评估"),
    ("8. Dify / Workflow 对照", "理论基础篇", 35, 35, "理解 Dify 如何抽象模型、知识库、工作流和 Agent"),
    ("8. Dify / Workflow 对照", "进阶篇", 36, 36, "把代码版 RAG 映射到 Dify 知识库和配置"),
    ("8. Dify / Workflow 对照", "实战篇", 37, 37, "完成 Dify Workflow / Agent 对照实验"),
    ("9. Agent 工程 / OpenClaw / AI Coding / 毕业项目", "理论基础篇", 38, 41, "理解 Agent 设计模式、AI Coding 上下文、OpenClaw 风险和产品边界"),
    ("9. Agent 工程 / OpenClaw / AI Coding / 毕业项目", "进阶篇", 42, 43, "补齐产品化、安全、评估和阶段复盘"),
    ("9. Agent 工程 / OpenClaw / AI Coding / 毕业项目", "实战篇", 44, 50, "完成毕业项目：知识库 + 工具 + Memory + 评估 + Dify 对照"),
    ("10. 微调认知 / 数据集 / 训练 / 评估", "理论基础篇", 51, 52, "理解微调、SFT、LoRA、QLoRA 的真实用途和工程取舍"),
    ("10. 微调认知 / 数据集 / 训练 / 评估", "进阶篇", 53, 57, "构造指令数据、清洗数据、建立评估集、选择小模型和环境"),
    ("10. 微调认知 / 数据集 / 训练 / 评估", "实战篇", 58, 65, "跑 LoRA dry-run、加载 adapter、评估前后效果，并做技术取舍"),
    ("11. 微调工程化 / 综合复盘", "理论基础篇", 66, 67, "记录模型/数据/实验版本，并理解推理部署成本"),
    ("11. 微调工程化 / 综合复盘", "进阶篇", 68, 69, "把微调判断接入毕业项目，完成 PoC 或替代方案"),
    ("11. 微调工程化 / 综合复盘", "实战篇", 70, 70, "整理 70 天输出物，形成自己的技术地图"),
]


def parse_day_number(value: str) -> int:
    match = re.search(r"\d+", str(value))
    if not match:
        raise ValueError(f"Cannot parse day number from {value!r}")
    return int(match.group(0))


def parse_resources_from_text(text: str) -> list[dict[str, str]]:
    resources: list[dict[str, str]] = []
    current_type = "other"
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("视频资源"):
            current_type = "video"
            continue
        if line.startswith("文档") or "GitHub" in line or "工具资源" in line:
            current_type = "doc"
            continue
        match = URL_RE.search(line)
        if not match:
            continue
        url = match.group(1)
        title_part = line[: match.start()].strip()
        title_part = re.sub(r"^\d+[、.]\s*", "", title_part)
        title = title_part.rstrip("：:").strip() or url
        resources.append({"title": title, "url": url, "resource_type": current_type})
    return resources


def module_name_for_day(day_number: int) -> str:
    for name, start, end in MODULE_RANGES:
        if start <= day_number <= end:
            return name
    raise ValueError(f"No module range for day {day_number}")


def create_schema() -> None:
    Base.metadata.create_all(engine)


def reset_tables() -> None:
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)


def find_plan_sheet(workbook):
    for name in ("70天按需打卡", "50天按需打卡"):
        if name in workbook.sheetnames:
            return workbook[name]
    raise ValueError("Workbook must contain 70天按需打卡 or 50天按需打卡")


def seed_quiz_questions(session) -> None:
    questions = [
        (1, "single", "知识更新类问题通常优先选择哪种方案？", ["Prompt", "RAG", "LoRA 微调", "只改 UI"], ["RAG"], "知识更新更适合通过检索增强生成处理，微调不适合作为知识库更新的默认方案。"),
        (3, "single", "Function Calling 中，真正执行工具的是谁？", ["模型", "应用程序", "用户浏览器", "视频平台"], ["应用程序"], "模型负责决定是否调用和给出参数，应用程序负责验证参数并执行工具。"),
        (7, "multiple", "RAG 质量通常受哪些因素影响？", ["chunk 策略", "检索召回", "rerank", "网页背景颜色"], ["chunk 策略", "检索召回", "rerank"], "RAG 的核心质量来自文档切分、召回、重排和答案生成链路。"),
        (10, "single", "LoRA 相比全量微调的主要优势是什么？", ["更省训练资源", "不需要数据", "一定更聪明", "能自动更新知识库"], ["更省训练资源"], "LoRA 通过训练少量 adapter 参数降低成本，但不代表一定更聪明。"),
    ]
    for module_sort_order, question_type, prompt, options, correct, explanation in questions:
        module = session.scalar(select(Module).where(Module.sort_order == module_sort_order))
        session.add(
            QuizQuestion(
                day_id=None,
                module_id=module.id if module else None,
                question_type=question_type,
                prompt=prompt,
                options_json=json.dumps(options, ensure_ascii=False),
                correct_answer_json=json.dumps(correct, ensure_ascii=False),
                explanation=explanation,
            )
        )


def seed_from_workbook(workbook_path: Path) -> None:
    reset_tables()
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = find_plan_sheet(workbook)

    with SessionLocal() as session:
        modules: dict[str, Module] = {}
        for sort_order, (name, _start, _end) in enumerate(MODULE_RANGES, start=1):
            module = Module(name=name, sort_order=sort_order, description=None)
            session.add(module)
            modules[name] = module
        session.flush()

        chapters: list[Chapter] = []
        for module_name, chapter_name, start, end, goal in CHAPTERS:
            chapter = Chapter(
                module_id=modules[module_name].id,
                name=chapter_name,
                start_day=start,
                end_day=end,
                goal=goal,
            )
            session.add(chapter)
            chapters.append(chapter)
        session.flush()

        def chapter_for_day(day_number: int) -> Chapter | None:
            for chapter in chapters:
                if chapter.start_day <= day_number <= chapter.end_day:
                    return chapter
            return None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            day_number = parse_day_number(row[0])
            module = modules[module_name_for_day(day_number)]
            chapter = chapter_for_day(day_number)
            learning_day = LearningDay(
                day_number=day_number,
                module_id=module.id,
                chapter_id=chapter.id if chapter else None,
                stage=str(row[1] or ""),
                topic=str(row[2] or ""),
                environment=str(row[3] or ""),
                learning_goal=str(row[4] or ""),
                resources_text=str(row[5] or ""),
                resource_hint=str(row[6] or ""),
                practice_steps=str(row[7] or ""),
                acceptance_criteria=str(row[8] or ""),
                output_artifact=str(row[9] or ""),
                status=str(row[10] or "未开始"),
                guidance=str(row[11] or ""),
            )
            session.add(learning_day)
            session.flush()

            for resource in parse_resources_from_text(learning_day.resources_text):
                session.add(
                    Resource(
                        day_id=learning_day.id,
                        module_id=module.id,
                        title=resource["title"],
                        url=resource["url"],
                        resource_type=resource["resource_type"],
                        notes=None,
                    )
                )

        seed_quiz_questions(session)
        ensure_resource_catalog(session)
        session.commit()


def main() -> None:
    path = resolve_project_path(SOURCE_WORKBOOK)
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")
    create_schema()
    seed_from_workbook(path)
    print(f"Seeded database from {path}")


if __name__ == "__main__":
    main()
